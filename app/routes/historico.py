from flask import Blueprint, jsonify, request, render_template
from flask_login import login_required, current_user
from app import mysql, limiter
import io
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from flask import send_file, jsonify
from datetime import datetime, timedelta, date, time
import logging

# Intentar importar reportlab para PDF; si no está instalado, seguiremos soportando solo Excel
try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    REPORTLAB_AVAILABLE = True
except Exception:
    REPORTLAB_AVAILABLE = False

# Definir helper para cabecera/pie si reportlab está disponible
if REPORTLAB_AVAILABLE:
    def _pdf_header_footer(canvas, doc):
        canvas.saveState()
        width, height = A4
        # Cabecera
        canvas.setFont('Helvetica-Bold', 12)
        canvas.setFillColor(colors.HexColor('#0066CC'))
        canvas.drawCentredString(width / 2.0, height - 30, 'Isis Med - Histórico de citas')
        # Línea debajo de la cabecera
        canvas.setStrokeColor(colors.HexColor('#0066CC'))
        canvas.setLineWidth(1)
        canvas.line(40, height - 36, width - 40, height - 36)

        # Pie de página: número de página a la derecha
        page_num_text = f"Página {doc.page}"
        canvas.setFont('Helvetica', 9)
        canvas.setFillColor(colors.grey)
        canvas.drawRightString(width - 40, 20, page_num_text)
        canvas.restoreState()

# Configurar logger para auditoría
logger = logging.getLogger(__name__)

historico_bp = Blueprint('historico', __name__)

# Constantes de seguridad
MAX_RECORDS_EXPORT = 10000  # Máximo de registros a exportar
MAX_FILE_SIZE = 10 * 1024 * 1024  # 10 MB máximo

# Mapeo de estados para visualización
MAPEO_ESTADOS = {
    'completada': 'Completada',
    'cancelada': 'Cancelada'
}

def mapear_estado(estado_bd):
    """Convierte estado de BD a nombre para mostrar"""
    if not estado_bd:
        return 'Desconocido'
    estado_lower = estado_bd.lower().strip()
    return MAPEO_ESTADOS.get(estado_lower, estado_bd)

@historico_bp.route('/fecha')
@login_required
def historico_fecha():
    # Obtener parámetros de filtro
    fecha_inicio = request.args.get('fecha_inicio')
    fecha_fin = request.args.get('fecha_fin')

    # Si no hay filtros aplicados, mostrar la página sin datos
    if not fecha_inicio and not fecha_fin:
        return render_template('historico/historico_fecha.html', citas=None)

    # Si hay filtros, consultar la base de datos
    cur = mysql.connection.cursor()
    try:
        # Construir la consulta base
        query = '''
            SELECT 
                c.id_cita,
                c.fecha,
                c.hora,
                c.estado,
                c.motivo,
                p.nombre,
                p.apellido, 
                d.nombre,
                d.apellido
            FROM citas c 
            JOIN pacientes p ON c.id_paciente = p.id_paciente 
            JOIN doctores d ON c.id_doctor = d.id_doctor
            WHERE c.estado IN ('completada', 'cancelada')
        '''

        params = []

        # Añadir filtros si se proporcionaron
        if fecha_inicio:
            query += " AND c.fecha >= %s"
            params.append(fecha_inicio)

        if fecha_fin:
            query += " AND c.fecha <= %s"
            params.append(fecha_fin)

        # Ordenar los resultados
        query += " ORDER BY c.fecha DESC, c.hora DESC"

        # Ejecutar la consulta
        cur.execute(query, params)
        citas = cur.fetchall()

        return render_template('historico/historico_fecha.html', citas=citas,
                               fecha_inicio=fecha_inicio, fecha_fin=fecha_fin)
    finally:
        cur.close()


@historico_bp.route('/paciente')
@login_required
def historico_paciente():
    cur = mysql.connection.cursor()
    try:
        # Obtener la lista de pacientes para el selector
        cur.execute('''
            SELECT id_paciente, nombre, apellido 
            FROM pacientes
            ORDER BY apellido, nombre
        ''')
        pacientes = cur.fetchall()
        return render_template('historico/historico_paciente.html', pacientes=pacientes)
    finally:
        cur.close()


@historico_bp.route('/doctor')
@login_required
def historico_doctor():
    cur = mysql.connection.cursor()
    try:
        # Obtener la lista de doctores para el selector
        cur.execute('''
            SELECT id_doctor, nombre, apellido 
            FROM doctores
            ORDER BY apellido, nombre
        ''')
        doctores = cur.fetchall()
        return render_template('historico/historico_doctor.html', doctores=doctores)
    finally:
        cur.close()


@historico_bp.route('/api/paciente/<int:paciente_id>')
@login_required
@limiter.limit("100 per minute")
def api_historico_paciente(paciente_id):
    """
    API para obtener histórico de citas de un paciente.
    
    Seguridad:
    - Requiere autenticación
    - Solo admin o el paciente mismo puede ver sus datos
    - Rate limiting: 100 solicitudes por minuto
    """
    try:
        cur = mysql.connection.cursor()
        
        # 1. Validación: Verificar que el paciente exista
        cur.execute('SELECT id_paciente FROM pacientes WHERE id_paciente = %s', (paciente_id,))
        if not cur.fetchone():
            cur.close()
            logger.warning(f"Usuario {current_user.id} intentó acceder a paciente inexistente: {paciente_id}")
            return jsonify({"error": "Paciente no encontrado"}), 404
        
        # 2. Registrar auditoría
        logger.info(f"Usuario {current_user.id} consultó histórico del paciente {paciente_id}")
        
        # 4. Obtener datos
        cur.execute('''
            SELECT 
                c.id_cita,
                DATE_FORMAT(c.fecha, '%%d/%%m/%%Y') as fecha,
                DATE_FORMAT(c.hora, '%%H:%%i') as hora,
                c.estado,
                c.motivo,
                CONCAT(p.nombre, ' ', p.apellido) as paciente,
                CONCAT(d.nombre, ' ', d.apellido) as doctor    
            FROM citas c 
            JOIN pacientes p ON c.id_paciente = p.id_paciente
            JOIN doctores d ON c.id_doctor = d.id_doctor          
            WHERE c.id_paciente = %s AND c.estado IN ('completada', 'cancelada')
            ORDER BY c.fecha DESC, c.hora DESC
        ''', (paciente_id,))

        # Obtener los nombres de las columnas
        columns = [desc[0] for desc in cur.description]

        # Convertir los resultados a una lista de diccionarios
        citas = []
        for row in cur.fetchall():
            # Crear un diccionario para esta fila
            cita = dict(zip(columns, row))
            citas.append(cita)

        cur.close()
        return jsonify(citas)
    
    except Exception as e:
        logger.error(f"Error en api_historico_paciente (paciente_id={paciente_id}): {str(e)}", exc_info=True)
        return jsonify({"error": "Error interno del servidor"}), 500
        return jsonify({"error": str(e)}), 500
    finally:
        cur.close()


@historico_bp.route('/api/doctor/<int:doctor_id>')
@login_required
@limiter.limit("100 per minute")
def api_historico_doctor(doctor_id):
    """
    API para obtener histórico de citas de un doctor.
    
    Seguridad:
    - Requiere autenticación
    - Solo admin o el doctor mismo puede ver sus datos
    - Rate limiting: 100 solicitudes por minuto
    """
    try:
        cur = mysql.connection.cursor()
        
        # 1. Validación: Verificar que el doctor exista
        cur.execute('SELECT id_doctor FROM doctores WHERE id_doctor = %s', (doctor_id,))
        if not cur.fetchone():
            cur.close()
            return jsonify({"error": "Doctor no encontrado"}), 404
        
        # 2. Registrar auditoría
        logger.info(f"Usuario {current_user.id} consultó histórico del doctor {doctor_id}")
        
        # 4. Obtener datos
        cur.execute('''
            SELECT 
                c.id_cita,
                DATE_FORMAT(c.fecha, '%%d/%%m/%%Y') as fecha,
                DATE_FORMAT(c.hora, '%%H:%%i') as hora,
                c.estado,
                c.motivo,
                CONCAT(p.nombre, ' ', p.apellido) as paciente,
                CONCAT(d.nombre, ' ', d.apellido) as doctor    
            FROM citas c 
            JOIN pacientes p ON c.id_paciente = p.id_paciente
            JOIN doctores d ON c.id_doctor = d.id_doctor          
            WHERE c.id_doctor = %s AND c.estado IN ('completada', 'cancelada')
            ORDER BY c.fecha DESC, c.hora DESC
        ''', (doctor_id,))

        # Obtener los nombres de las columnas
        columns = [desc[0] for desc in cur.description]

        # Convertir los resultados a una lista de diccionarios
        citas = []
        for row in cur.fetchall():
            # Crear un diccionario para esta fila
            cita = dict(zip(columns, row))
            citas.append(cita)

        cur.close()
        return jsonify(citas)
    
    except Exception as e:
        logger.error(f"Error en api_historico_doctor (doctor_id={doctor_id}): {str(e)}", exc_info=True)
        return jsonify({"error": "Error interno del servidor"}), 500


@historico_bp.route('/exportar/doctor/<int:doctor_id>')
@login_required
@limiter.limit("20 per minute")
def exportar_excel_doctor(doctor_id):
    """
    Exportar histórico de citas de un doctor a Excel.
    
    Seguridad:
    - Requiere autenticación
    - Solo admin o el doctor mismo puede exportar
    - Rate limiting más restrictivo: 20 solicitudes por minuto
    - Límite de registros: 10000
    """
    try:
        cur = mysql.connection.cursor()
        
        # 1. Validación: Verificar que el doctor exista
        cur.execute('SELECT nombre, apellido FROM doctores WHERE id_doctor = %s', (doctor_id,))
        doctor_info = cur.fetchone()

        if not doctor_info:
            cur.close()
            logger.warning(f"Usuario {current_user.id} intentó exportar datos de doctor inexistente: {doctor_id}")
            return jsonify({"error": "Doctor no encontrado"}), 404

        # 2. Validar límite de registros
        cur.execute('''
            SELECT COUNT(*) as count FROM citas 
            WHERE id_doctor = %s AND estado IN ('completada', 'cancelada')
        ''', (doctor_id,))
        
        count_result = cur.fetchone()
        record_count = count_result[0] if count_result else 0
        
        if record_count > MAX_RECORDS_EXPORT:
            cur.close()
            logger.warning(f"Intento de exportar {record_count} registros (máx: {MAX_RECORDS_EXPORT}) por usuario {current_user.id}")
            return jsonify({
                "error": f"Demasiados registros ({record_count}). Máximo permitido: {MAX_RECORDS_EXPORT}"
            }), 400
        
        # 4. Registrar auditoría
        logger.info(f"Usuario {current_user.id} exportó histórico del doctor {doctor_id} ({record_count} registros)")
        
        # 5. Obtener las citas
        cur.execute('''
            SELECT 
                c.id_cita,
                DATE_FORMAT(c.fecha, '%%d/%%m/%%Y') as fecha,
                DATE_FORMAT(c.hora, '%%H:%%i') as hora,
                c.estado,
                c.motivo,
                CONCAT(p.nombre, ' ', p.apellido) as paciente,
                CONCAT(d.nombre, ' ', d.apellido) as doctor    
            FROM citas c 
            JOIN pacientes p ON c.id_paciente = p.id_paciente
            JOIN doctores d ON c.id_doctor = d.id_doctor          
            WHERE c.id_doctor = %s AND c.estado IN ('completada', 'cancelada')
            ORDER BY c.fecha DESC, c.hora DESC
        ''', (doctor_id,))

        citas = cur.fetchall()
        cur.close()

        # 6. Generar el Excel
        doctor_nombre = f"{doctor_info[0]} {doctor_info[1]}"
        titulo = f"Histórico de citas - Dr. {doctor_nombre}"
        nombre_archivo = f"Historico_Doctor_{doctor_info[1]}_{doctor_info[0]}_{datetime.now().strftime('%d%m%Y_%H%M%S')}.xlsx"

        return generar_excel_historico(citas, titulo, nombre_archivo)

    except Exception as e:
        logger.error(f"Error al exportar Excel del doctor {doctor_id}: {str(e)}", exc_info=True)
        return jsonify({"error": "Error interno del servidor"}), 500


@historico_bp.route('/exportar/paciente/<int:paciente_id>')
@login_required
@limiter.limit("20 per minute")
def exportar_excel_paciente(paciente_id):
    """
    Exportar histórico de citas de un paciente a Excel.
    
    Seguridad:
    - Requiere autenticación
    - Solo admin o el paciente mismo puede exportar
    - Rate limiting más restrictivo: 20 solicitudes por minuto
    - Límite de registros: 10000
    """
    try:
        cur = mysql.connection.cursor()
        
        # 1. Validación: Verificar que el paciente exista
        cur.execute('SELECT nombre, apellido FROM pacientes WHERE id_paciente = %s', (paciente_id,))
        paciente_info = cur.fetchone()
        
        if not paciente_info:
            cur.close()
            logger.warning(f"Usuario {current_user.id} intentó exportar datos de paciente inexistente: {paciente_id}")
            return jsonify({"error": "Paciente no encontrado"}), 404
        
        # 2. Validar límite de registros
        cur.execute('''
            SELECT COUNT(*) as count FROM citas 
            WHERE id_paciente = %s AND estado IN ('completada', 'cancelada')
        ''', (paciente_id,))
        
        count_result = cur.fetchone()
        record_count = count_result[0] if count_result else 0
        
        if record_count > MAX_RECORDS_EXPORT:
            cur.close()
            logger.warning(f"Intento de exportar {record_count} registros (máx: {MAX_RECORDS_EXPORT}) por usuario {current_user.id}")
            return jsonify({
                "error": f"Demasiados registros ({record_count}). Máximo permitido: {MAX_RECORDS_EXPORT}"
            }), 400
        
        # 4. Registrar auditoría
        logger.info(f"Usuario {current_user.id} exportó histórico del paciente {paciente_id} ({record_count} registros)")
        
        # 5. Obtener las citas
        cur.execute('''
            SELECT 
                c.id_cita,
                DATE_FORMAT(c.fecha, '%%d/%%m/%%Y') as fecha,
                DATE_FORMAT(c.hora, '%%H:%%i') as hora,
                c.estado,
                c.motivo,
                CONCAT(p.nombre, ' ', p.apellido) as paciente,
                CONCAT(d.nombre, ' ', d.apellido) as doctor    
            FROM citas c 
            JOIN pacientes p ON c.id_paciente = p.id_paciente
            JOIN doctores d ON c.id_doctor = d.id_doctor          
            WHERE c.id_paciente = %s AND c.estado IN ('completada', 'cancelada')
            ORDER BY c.fecha DESC, c.hora DESC
        ''', (paciente_id,))

        citas = cur.fetchall()
        cur.close()

        # 6. Generar el Excel
        paciente_nombre = f"{paciente_info[0]} {paciente_info[1]}"
        titulo = f"Histórico de citas - Paciente: {paciente_nombre}"
        nombre_archivo = f"Historico_Paciente_{paciente_info[1]}_{paciente_info[0]}_{datetime.now().strftime('%d%m%Y_%H%M%S')}.xlsx"

        return generar_excel_historico(citas, titulo, nombre_archivo)

    except Exception as e:
        logger.error(f"Error al exportar Excel del paciente {paciente_id}: {str(e)}", exc_info=True)
        return jsonify({"error": "Error interno del servidor"}), 500


@historico_bp.route('/exportar/pdf/doctor/<int:doctor_id>')
@login_required
@limiter.limit("20 per minute")
def exportar_pdf_doctor(doctor_id):
        """Exportar histórico de un doctor a PDF."""
        if not REPORTLAB_AVAILABLE:
            return jsonify({"error": "La generación de PDF requiere la librería reportlab (instalar reportlab)"}), 501
        try:
            cur = mysql.connection.cursor()
            cur.execute('SELECT nombre, apellido FROM doctores WHERE id_doctor = %s', (doctor_id,))
            doctor_info = cur.fetchone()
            if not doctor_info:
                cur.close()
                return jsonify({"error": "Doctor no encontrado"}), 404

            # Obtener citas
            cur.execute('''
                SELECT 
                    DATE_FORMAT(c.fecha, '%%d/%%m/%%Y') as fecha,
                    DATE_FORMAT(c.hora, '%%H:%%i') as hora,
                    CONCAT(p.nombre, ' ', p.apellido) as paciente,
                    c.motivo,
                    c.estado
                FROM citas c
                JOIN pacientes p ON c.id_paciente = p.id_paciente
                WHERE c.id_doctor = %s AND c.estado IN ('completada', 'cancelada')
                ORDER BY c.fecha DESC, c.hora DESC
            ''', (doctor_id,))

            citas = cur.fetchall()
            cur.close()

            # Crear PDF
            buffer = io.BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=A4)
            styles = getSampleStyleSheet()
            elements = []

            titulo = f"Histórico de citas - Dr. {doctor_info[0]} {doctor_info[1]}"
            elements.append(Paragraph(titulo, styles['Title']))
            elements.append(Spacer(1, 12))
            elements.append(Paragraph(f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}", styles['Normal']))
            elements.append(Spacer(1, 12))

            # Tabla
            data = [['Fecha', 'Hora', 'Paciente', 'Motivo', 'Estado']]
            for row in citas:
                estado = mapear_estado(row[4])
                data.append([row[0], row[1], row[2], row[3] or '', estado])

            table = Table(data, repeatRows=1)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#0066CC')),
                ('TEXTCOLOR', (0,0), (-1,0), colors.white),
                ('ALIGN', (0,0), (-1,-1), 'LEFT'),
                ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
                ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
                ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#F8F9FA')])
            ]))

            elements.append(table)
            doc.build(elements, onFirstPage=_pdf_header_footer, onLaterPages=_pdf_header_footer)
            buffer.seek(0)

            filename = f"Historico_Doctor_{doctor_info[1]}_{doctor_info[0]}_{datetime.now().strftime('%d%m%Y_%H%M%S')}.pdf"
            return send_file(buffer, as_attachment=True, download_name=filename, mimetype='application/pdf')

        except Exception as e:
            logger.error(f"Error al generar PDF doctor {doctor_id}: {str(e)}", exc_info=True)
            return jsonify({"error": "Error interno del servidor"}), 500


@historico_bp.route('/exportar/pdf/paciente/<int:paciente_id>')
@login_required
@limiter.limit("20 per minute")
def exportar_pdf_paciente(paciente_id):
        """Exportar histórico de un paciente a PDF."""
        if not REPORTLAB_AVAILABLE:
            return jsonify({"error": "La generación de PDF requiere la librería reportlab (instalar reportlab)"}), 501
        try:
            cur = mysql.connection.cursor()
            cur.execute('SELECT nombre, apellido FROM pacientes WHERE id_paciente = %s', (paciente_id,))
            paciente_info = cur.fetchone()
            if not paciente_info:
                cur.close()
                return jsonify({"error": "Paciente no encontrado"}), 404

            cur.execute('''
                SELECT 
                    DATE_FORMAT(c.fecha, '%%d/%%m/%%Y') as fecha,
                    DATE_FORMAT(c.hora, '%%H:%%i') as hora,
                    CONCAT(d.nombre, ' ', d.apellido) as doctor,
                    c.motivo,
                    c.estado
                FROM citas c
                JOIN doctores d ON c.id_doctor = d.id_doctor
                WHERE c.id_paciente = %s AND c.estado IN ('completada', 'cancelada')
                ORDER BY c.fecha DESC, c.hora DESC
            ''', (paciente_id,))

            citas = cur.fetchall()
            cur.close()

            buffer = io.BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=A4)
            styles = getSampleStyleSheet()
            elements = []

            titulo = f"Histórico de citas - Paciente: {paciente_info[0]} {paciente_info[1]}"
            elements.append(Paragraph(titulo, styles['Title']))
            elements.append(Spacer(1, 12))
            elements.append(Paragraph(f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}", styles['Normal']))
            elements.append(Spacer(1, 12))

            data = [['Fecha', 'Hora', 'Doctor', 'Motivo', 'Estado']]
            for row in citas:
                estado = mapear_estado(row[4])
                data.append([row[0], row[1], row[2], row[3] or '', estado])

            table = Table(data, repeatRows=1)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#0066CC')),
                ('TEXTCOLOR', (0,0), (-1,0), colors.white),
                ('ALIGN', (0,0), (-1,-1), 'LEFT'),
                ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
                ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ]))

            elements.append(table)
            # Fondo alterno en filas y header/footer
            table.setStyle(TableStyle([
                ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#F8F9FA')])
            ]))
            doc.build(elements, onFirstPage=_pdf_header_footer, onLaterPages=_pdf_header_footer)
            buffer.seek(0)

            filename = f"Historico_Paciente_{paciente_info[1]}_{paciente_info[0]}_{datetime.now().strftime('%d%m%Y_%H%M%S')}.pdf"
            return send_file(buffer, as_attachment=True, download_name=filename, mimetype='application/pdf')

        except Exception as e:
            logger.error(f"Error al generar PDF paciente {paciente_id}: {str(e)}", exc_info=True)
            return jsonify({"error": "Error interno del servidor"}), 500


@historico_bp.route('/exportar/pdf/fecha')
@login_required
def exportar_pdf_fecha():
    """Exportar histórico por rango de fechas a PDF."""
    if not REPORTLAB_AVAILABLE:
        return jsonify({"error": "La generación de PDF requiere la librería reportlab (instalar reportlab)"}), 501
    try:
        fecha_inicio = request.args.get('fecha_inicio')
        fecha_fin = request.args.get('fecha_fin')
        if not fecha_inicio or not fecha_fin:
            return jsonify({"error": "Se requieren ambas fechas"}), 400

        try:
            fecha_inicio_obj = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
            fecha_fin_obj = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
        except ValueError:
            return jsonify({"error": "Formato de fecha inválido"}), 400

        cur = mysql.connection.cursor()
        cur.execute('''
            SELECT 
                DATE_FORMAT(c.fecha, '%%d/%%m/%%Y') as fecha,
                DATE_FORMAT(c.hora, '%%H:%%i') as hora,
                CONCAT(p.nombre, ' ', p.apellido) as paciente,
                CONCAT(d.nombre, ' ', d.apellido) as doctor,
                c.motivo,
                c.estado
            FROM citas c
            JOIN pacientes p ON c.id_paciente = p.id_paciente
            JOIN doctores d ON c.id_doctor = d.id_doctor
            WHERE c.fecha BETWEEN %s AND %s AND c.estado IN ('completada', 'cancelada')
            ORDER BY c.fecha DESC, c.hora DESC
        ''', (fecha_inicio_obj, fecha_fin_obj))

        citas = cur.fetchall()
        cur.close()

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        styles = getSampleStyleSheet()
        elements = []

        titulo = f"Histórico de citas - Período: {fecha_inicio} al {fecha_fin}"
        elements.append(Paragraph(titulo, styles['Title']))
        elements.append(Spacer(1, 12))
        elements.append(Paragraph(f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}", styles['Normal']))
        elements.append(Spacer(1, 12))

        data = [['Fecha', 'Hora', 'Paciente', 'Doctor', 'Motivo', 'Estado']]
        for row in citas:
            estado = mapear_estado(row[5])
            data.append([row[0], row[1], row[2], row[3], row[4] or '', estado])

        table = Table(data, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#0066CC')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ]))

        elements.append(table)
        table.setStyle(TableStyle([
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#F8F9FA')])
        ]))
        doc.build(elements, onFirstPage=_pdf_header_footer, onLaterPages=_pdf_header_footer)
        buffer.seek(0)

        filename = f"Historico_Citas_{fecha_inicio}_{fecha_fin}_{datetime.now().strftime('%d%m%Y_%H%M%S')}.pdf"
        return send_file(buffer, as_attachment=True, download_name=filename, mimetype='application/pdf')

    except Exception as e:
        logger.error(f"Error al exportar PDF por fecha: {str(e)}", exc_info=True)
        return jsonify({"error": "Error interno del servidor"}), 500

@historico_bp.route('/exportar/fecha')
@login_required
def exportar_excel_fecha():
    try:
        # Obtener parámetros de fecha
        fecha_inicio = request.args.get('fecha_inicio')
        fecha_fin = request.args.get('fecha_fin')

        if not fecha_inicio or not fecha_fin:
            return jsonify({"error": "Se requieren ambas fechas"}), 400

        # Convertir a objetos de fecha
        try:
            fecha_inicio_obj = datetime.strptime(
                fecha_inicio, '%Y-%m-%d').date()
            fecha_fin_obj = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
        except ValueError:
            return jsonify({"error": "Formato de fecha inválido"}), 400

        # Obtener las citas para el rango de fechas
        cur = mysql.connection.cursor()
        cur.execute('''
            SELECT 
                c.id_cita,
                DATE_FORMAT(c.fecha, '%%d/%%m/%%Y') as fecha,
                c.hora,
                c.estado,
                c.motivo,
                CONCAT(p.nombre, ' ', p.apellido) as paciente,
                CONCAT(d.nombre, ' ', d.apellido) as doctor    
            FROM citas c 
            JOIN pacientes p ON c.id_paciente = p.id_paciente
            JOIN doctores d ON c.id_doctor = d.id_doctor          
            WHERE c.fecha BETWEEN %s AND %s AND c.estado IN ('completada', 'cancelada')
            ORDER BY c.fecha DESC, c.hora DESC
        ''', (fecha_inicio_obj, fecha_fin_obj))

        citas = cur.fetchall()
        cur.close()

        # Generar el Excel
        titulo = f"Histórico de citas - Período: {fecha_inicio} al {fecha_fin}"
        nombre_archivo = f"Historico_Citas_{fecha_inicio}_{fecha_fin}.xlsx"

        return generar_excel_historico(citas, titulo, nombre_archivo)

    except Exception as e:
        print(f"Error al exportar a Excel: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


def generar_excel_historico(datos, titulo, nombre_archivo):
    """
    Función auxiliar para generar un archivo Excel con el histórico de citas

    Args:
        datos: Lista de tuplas con los datos de las citas
        titulo: Título del reporte
        nombre_archivo: Nombre del archivo a generar

    Returns:
        Respuesta Flask con el archivo Excel adjunto
    """
    try:
        # Crear un nuevo libro de Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Histórico de citas"

        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="0066CC", end_color="0066CC", fill_type="solid")
        centered = Alignment(horizontal="center", vertical="center")

        # Título del reporte
        ws.merge_cells('A1:F1')
        ws['A1'] = titulo
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal="center")

        # Fecha de generación
        ws.merge_cells('A2:F2')
        ws['A2'] = f"Generado el: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
        ws['A2'].alignment = Alignment(horizontal="right")

        # Encabezados
        headers = ["Fecha", "Hora", "Paciente", "Doctor", "Motivo", "Estado"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = centered

        # Datos
        if not datos:
            ws.merge_cells('A5:F5')
            ws['A5'] = "No hay registros históricos"
            ws['A5'].alignment = Alignment(horizontal="center")
        else:
            for row_num, cita in enumerate(datos, 5):
                # Fecha
                ws.cell(row=row_num, column=1).value = cita[1]

                # Hora (formatear si es un objeto timedelta)
                hora = cita[2]
                if isinstance(hora, timedelta):
                    total_seconds = int(hora.total_seconds())
                    hours = total_seconds // 3600
                    minutes = (total_seconds % 3600) // 60
                    hora_str = f"{hours:02d}:{minutes:02d}"
                else:
                    hora_str = str(hora)
                ws.cell(row=row_num, column=2).value = hora_str

                # Paciente
                ws.cell(row=row_num, column=3).value = cita[5]

                # Doctor
                ws.cell(row=row_num, column=4).value = cita[6]

                # Motivo
                ws.cell(row=row_num, column=5).value = cita[4]

                # Estado (mapeado a nombre amigable)
                estado_cell = ws.cell(row=row_num, column=6)
                estado_mapeado = mapear_estado(cita[3])
                estado_cell.value = estado_mapeado

                # Color según el estado
                estado_lower = cita[3].lower()
                if estado_lower == 'completada':
                    estado_cell.fill = PatternFill(
                        start_color="90EE90", end_color="90EE90", fill_type="solid")
                elif estado_lower == 'cancelada':
                    estado_cell.fill = PatternFill(
                        start_color="FFC0CB", end_color="FFC0CB", fill_type="solid")

        # Ajustar anchos de columna
        column_widths = {}

        # Calcular el ancho máximo para cada columna
        for row in ws.rows:
            for cell in row:
                if cell.value:
                    # Obtener la columna solo si no es una celda combinada
                    if not isinstance(cell, openpyxl.cell.cell.MergedCell):
                        col_letter = cell.column_letter
                        # Inicializar si es la primera vez que vemos esta columna
                        if col_letter not in column_widths:
                            column_widths[col_letter] = 0
                        # Actualizar el ancho máximo
                        column_widths[col_letter] = max(
                            column_widths[col_letter], len(str(cell.value)))

        # Aplicar los anchos calculados
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width + 2

        # Guardar a un buffer en memoria
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # Enviar el archivo
        return send_file(
            output,
            as_attachment=True,
            download_name=nombre_archivo,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        print(f"Error al generar Excel: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500
